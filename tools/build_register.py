#!/usr/bin/env python3
"""
Kétirányú Markdown <-> XLSX konverter a teszteset-regiszterhez.

A regiszter két alakban létezik: `docs/tesztesetek_v<n>.md` az olvasható forrás, az azonos
nevű `.xlsx` pedig az olvasható, Excelben megnyitható alak. A
kettőt eddig kézzel tartottuk szinkronban. Ez ugyanaz a kétforrású alakzat, amit a plugin
kódjában már hatszor megtaláltunk -- csak a dokumentációra alkalmazva.

Innentől a markdown az egyetlen forrás, az xlsx belőle generálódik, és a `verify` minden
építés után lefuttatja a md -> xlsx -> md' kört. Eltérés esetén nem írunk fájlt.

A markdown alakja (a v33 óta változatlan):

    ## Lista oldal (77 teszteset)

    ### TC-List-001 — List-001 (Happy path) [Funkcionális (kézi)]
    **Előfeltétel:** ...
    **Lépések:** 1. ...
    2. ...
    **Elvárt:** ...

A fejlécben szereplő darabszám GENERÁLT: a konverter megszámolja, nem átmásolja. A v33-ban ez
a szám kilenc szekcióból hétben elavult volt -- a v29-es állapotot hirdette, négy kiadáson át,
mert egy kézzel karbantartott összeg némán romlik el.

A követelménymező tartalmazhat zárójelet (pl. `Req-006 (Seed)`), ezért a Típus mindig az
UTOLSÓ zárójeles csoport a szögletes zárójel előtt. Nem mohó illesztéssel ez az egy eset
csendben elcsúszna.

Használat:
    python3 build_register.py build   <bemenet.md> [-o kimenet.xlsx]
    python3 build_register.py reverse <bemenet.xlsx> [-o kimenet.md]
    python3 build_register.py verify  <bemenet.md>
    python3 build_register.py compare <bemenet.md> <meglévő.xlsx>

Függőség: openpyxl  (pip install openpyxl)
"""

import argparse
import collections
import os
import re
import shutil
import sys
import tempfile

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:
    sys.exit("HIÁNYZÓ FÜGGŐSÉG: pip install openpyxl")

COLUMNS = ["Teszteset ID", "Követelmény", "Típus", "Automatizálás típusa",
           "Előfeltétel", "Tesztlépések", "Elvárt eredmény"]
FIELDS = ["id", "req", "tipus", "aut", "elo", "lep", "elv"]
WIDTHS = {"A": 15.0, "B": 13.0, "C": 11.0, "D": 24.0, "E": 50.0, "F": 45.0}
ROW_FILL = {"Happy path": "FFEAF3DE", "Negatív": "FFFCEBEB", "Vizuális": "FFFFF8E6"}
HEADER_FILL = "FF4472C4"

SEC_RE = re.compile(r"^## (.+?) \((\d+) teszteset\)$")
# A Típus az UTOLSÓ zárójeles csoport: a követelmény maga is tartalmazhat zárójelet.
HDR_RE = re.compile(r"^### (TC-[A-Za-z0-9]+-\d+) — (.+) \(([^()]+)\) \[([^\[\]]+)\]$")
LABELS = (("elo", "**Előfeltétel:**"), ("lep", "**Lépések:**"), ("elv", "**Elvárt:**"))


class ContractError(Exception):
    """A regiszter olyat tartalmaz, amit a konverter nem ismer."""


def parse_md(md_path):
    lines = open(md_path, encoding="utf-8").read().split("\n")
    title = lines[0] if lines and lines[0].startswith("# ") else None
    sections, current, i = collections.OrderedDict(), None, 0

    while i < len(lines):
        line = lines[i]
        sec = SEC_RE.match(line)
        if sec:
            current = sec.group(1)
            if current in sections:
                raise ContractError(f"{md_path}:{i+1}: kétszer szereplő szekció: {current!r}")
            sections[current] = []
            i += 1
            continue

        if line.startswith("### "):
            if current is None:
                raise ContractError(f"{md_path}:{i+1}: teszteset szekciócím előtt")
            head = HDR_RE.match(line)
            if not head:
                raise ContractError(f"{md_path}:{i+1}: nem értelmezhető fejléc: {line[:90]!r}")
            rec = dict(zip(FIELDS[:4], head.groups()))
            rec.update({"elo": "", "lep": "", "elv": ""})
            j, field = i + 1, None
            while j < len(lines) and not lines[j].startswith(("### ", "## ")):
                text = lines[j]
                for key, prefix in LABELS:
                    if text.startswith(prefix):
                        field = key
                        rec[key] = text[len(prefix):].strip()
                        break
                else:
                    if text.strip() and field:
                        rec[field] += "\n" + text.strip()
                j += 1
            sections[current].append(rec)
            i = j
            continue
        i += 1

    if not sections:
        raise ContractError(f"{md_path}: egyetlen szekció sem található")
    return title, sections


def md_to_xlsx(md_path, xlsx_path):
    _, sections = parse_md(md_path)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for name, rows in sections.items():
        ws = wb.create_sheet(title=name)
        for col, label in enumerate(COLUMNS, start=1):
            cell = ws.cell(1, col, label)
            cell.font = Font(bold=True, color="FFFFFFFF")
            cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
            cell.alignment = Alignment(wrap_text=True)

        for n, rec in enumerate(rows, start=2):
            fill = ROW_FILL.get(rec["tipus"])
            if fill is None:
                raise ContractError(
                    f"ismeretlen Típus: {rec['tipus']!r} ({rec['id']}) -- "
                    f"a sorszínezés nem dönthető el; ismert: {', '.join(ROW_FILL)}")
            for col, key in enumerate(FIELDS, start=1):
                cell = ws.cell(n, col, rec[key])
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.fill = PatternFill("solid", fgColor=fill)

        for letter, width in WIDTHS.items():
            ws.column_dimensions[letter].width = width
        ws.freeze_panes = "A2"

    wb.save(xlsx_path)
    return xlsx_path


def xlsx_to_md(xlsx_path, md_path, title=None):
    wb = openpyxl.load_workbook(xlsx_path)
    out = [title or "# local_artqtml Tesztesetek", ""]

    for name in wb.sheetnames:
        ws = wb[name]
        header = [ws.cell(1, c).value for c in range(1, len(COLUMNS) + 1)]
        if header != COLUMNS:
            raise ContractError(f"{name}: váratlan fejléc: {header}")
        rows = []
        for r in range(2, ws.max_row + 1):
            vals = [ws.cell(r, c).value or "" for c in range(1, len(COLUMNS) + 1)]
            if not str(vals[0]).strip():
                continue
            rows.append(dict(zip(FIELDS, [str(v) for v in vals])))

        out.append(f"## {name} ({len(rows)} teszteset)")
        out.append("")
        for rec in rows:
            out.append(f"### {rec['id']} — {rec['req']} ({rec['tipus']}) [{rec['aut']}]")
            out.append(f"**Előfeltétel:** {rec['elo']}")
            out.append(f"**Lépések:** {rec['lep']}")
            out.append(f"**Elvárt:** {rec['elv']}")
            out.append("")
        out.append("")

    text = re.sub(r"\n{3,}", "\n\n", "\n".join(out)).strip() + "\n"
    open(md_path, "w", encoding="utf-8").write(text)
    return md_path


def canonicalise_md(md_path):
    """A darabszámokat és a blokk-elválasztást a parse-olt tartalomból írjuk vissza."""
    title, _ = parse_md(md_path)
    tmpdir = tempfile.mkdtemp(prefix="build_register_")
    try:
        tmp_xlsx = os.path.join(tmpdir, "r.xlsx")
        md_to_xlsx(md_path, tmp_xlsx)
        xlsx_to_md(tmp_xlsx, md_path, title)
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)


def verify(md_path):
    canonicalise_md(md_path)
    tmpdir = tempfile.mkdtemp(prefix="build_register_")
    try:
        tmp_xlsx, tmp_md = os.path.join(tmpdir, "r.xlsx"), os.path.join(tmpdir, "r.md")
        title, sections = parse_md(md_path)
        md_to_xlsx(md_path, tmp_xlsx)
        xlsx_to_md(tmp_xlsx, tmp_md, title)
        a = open(md_path, encoding="utf-8").read().strip()
        b = open(tmp_md, encoding="utf-8").read().strip()
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)

    total = sum(len(v) for v in sections.values())
    if a == b:
        print(f"KÖR RENDBEN: {os.path.basename(md_path)} -> xlsx -> markdown, "
              f"{total} teszteset {len(sections)} szekcióban, eltérés nincs")
        return True
    import difflib
    diff = list(difflib.unified_diff(a.split("\n"), b.split("\n"), "eredeti", "vissza", lineterm="", n=1))
    print("KÖR ELTÉRÉS:")
    print("\n".join(diff[:40]))
    return False


def compare(md_path, xlsx_path):
    """A generált és egy meglévő xlsx cellánkénti összevetése -- a hűség bizonyítéka."""
    tmpdir = tempfile.mkdtemp(prefix="build_register_")
    try:
        generated = md_to_xlsx(md_path, os.path.join(tmpdir, "gen.xlsx"))
        a, b = openpyxl.load_workbook(generated), openpyxl.load_workbook(xlsx_path)
        if a.sheetnames != b.sheetnames:
            print(f"munkalapok eltérnek:\n  generált: {a.sheetnames}\n  meglévő:  {b.sheetnames}")
            return False
        bad = cells = 0
        for name in a.sheetnames:
            sa, sb = a[name], b[name]
            for r in range(1, max(sa.max_row, sb.max_row) + 1):
                for c in range(1, len(COLUMNS) + 1):
                    cells += 1
                    va = (sa.cell(r, c).value or "")
                    vb = (sb.cell(r, c).value or "")
                    if str(va).strip() != str(vb).strip():
                        if bad < 5:
                            print(f"  {name} R{r}C{c}:\n    generált: {str(va)[:70]!r}\n"
                                  f"    meglévő:  {str(vb)[:70]!r}")
                        bad += 1
        print(f"cellák: {cells}, eltérés: {bad}")
        return bad == 0
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("command", choices=["build", "reverse", "verify", "compare"])
    ap.add_argument("input")
    ap.add_argument("second", nargs="?")
    ap.add_argument("-o", "--output")
    args = ap.parse_args()

    try:
        if args.command == "build":
            out = args.output or os.path.splitext(args.input)[0] + ".xlsx"
            canonicalise_md(args.input)
            md_to_xlsx(args.input, out)
            print(f"kész: {out}")
            if not verify(args.input):
                sys.exit("A kör-ellenőrzés elbukott -- az xlsx nem hű a markdownhoz.")
        elif args.command == "reverse":
            out = args.output or os.path.splitext(args.input)[0] + ".from-xlsx.md"
            xlsx_to_md(args.input, out)
            print(f"kész: {out}")
        elif args.command == "compare":
            if not args.second:
                sys.exit("a compare két argumentumot vár: <md> <xlsx>")
            sys.exit(0 if compare(args.input, args.second) else 1)
        else:
            sys.exit(0 if verify(args.input) else 1)
    except ContractError as exc:
        sys.exit(f"KÉSZLETEN KÍVÜLI SZERKEZET: {exc}")


if __name__ == "__main__":
    main()
